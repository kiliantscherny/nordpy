"""Tests for nordpy.export — CSV, XLSX, DuckDB export with model flattening."""

from __future__ import annotations

import csv

import pytest

from nordpy.export import (
    _get_headers,
    _get_rows,
    _model_to_flat_dict,
    _python_to_sql_type,
    export_csv,
    export_duckdb,
    export_xlsx,
)
from nordpy.models import (
    Holding,
    Transaction,
)


# ── _model_to_flat_dict ──


class TestModelToFlatDict:
    def test_flat_model(self):
        """Simple model with no nested BaseModel fields."""
        from nordpy.models import Account

        acc = Account(accid=1, accno="123", type="ASK")
        flat = _model_to_flat_dict(acc)
        assert flat["accid"] == 1
        assert flat["accno"] == "123"
        assert flat["type"] == "ASK"

    def test_nested_model(self):
        """Nested BaseModel fields get flattened with prefix.
        Must use model_validate — Holding's _parse_money needs dict input."""
        h = Holding.model_validate(
            {
                "instrument": {"name": "Apple", "symbol": "AAPL", "isin": "US123"},
                "qty": 10.0,
                "acq_price": {"value": 150.0, "currencyCode": "USD"},
                "market_value": {"value": 175.0, "currencyCode": "USD"},
            }
        )
        flat = _model_to_flat_dict(h)
        assert flat["instrument_name"] == "Apple"
        assert flat["instrument_symbol"] == "AAPL"
        assert flat["acq_price_value"] == 150.0
        assert flat["acq_price_currency"] == "USD"

    def test_date_field(self):
        """Date fields get converted to ISO format strings."""
        tx = Transaction.model_validate(
            {
                "transactionId": "tx-1",
                "accountingDate": "2024-06-15",
                "transactionTypeName": "BUY",
                "amount": {"value": -100.0},
            }
        )
        flat = _model_to_flat_dict(tx)
        assert flat["accounting_date"] == "2024-06-15"


# ── _get_headers / _get_rows ──


class TestHeadersAndRows:
    def test_empty_list(self):
        assert _get_headers([]) == []
        assert _get_rows([]) == []

    def test_populated(self):
        from nordpy.models import Account

        data = [
            Account(accid=1, accno="111", type="ASK"),
            Account(accid=2, accno="222", type="ISK"),
        ]
        headers = _get_headers(data)
        rows = _get_rows(data)
        assert "accid" in headers
        assert "accno" in headers
        assert len(rows) == 2
        assert rows[0][headers.index("accid")] == 1


# ── _python_to_sql_type ──


class TestPythonToSqlType:
    def test_float(self):
        assert _python_to_sql_type(3.14) == "DOUBLE"

    def test_int(self):
        assert _python_to_sql_type(42) == "BIGINT"

    def test_bool(self):
        """bool is a subclass of int in Python, so isinstance(True, int) is True.
        The function checks int before bool, so booleans map to BIGINT."""
        assert _python_to_sql_type(True) == "BIGINT"

    def test_str(self):
        assert _python_to_sql_type("hello") == "TEXT"

    def test_none(self):
        assert _python_to_sql_type(None) == "TEXT"


# ── CSV export ──


class TestExportCSV:
    def test_creates_file(self, tmp_path, monkeypatch):
        import nordpy.export as export_mod

        monkeypatch.setattr(export_mod, "EXPORT_DIR", tmp_path)

        from nordpy.models import Account

        data = [Account(accid=1, accno="123", type="ASK")]
        path = export_csv(data, "test_accounts")

        assert path.exists()
        assert path.suffix == ".csv"

    def test_correct_content(self, tmp_path, monkeypatch):
        import nordpy.export as export_mod

        monkeypatch.setattr(export_mod, "EXPORT_DIR", tmp_path)

        from nordpy.models import Account

        data = [
            Account(accid=1, accno="111", type="ASK"),
            Account(accid=2, accno="222", type="ISK"),
        ]
        path = export_csv(data, "accounts")

        with open(path, encoding="utf-8") as f:
            reader = csv.reader(f)
            rows = list(reader)

        assert len(rows) == 3  # header + 2 data rows
        assert "accid" in rows[0]
        assert "accno" in rows[0]


# ── XLSX export ──


class TestExportXLSX:
    def test_creates_file(self, tmp_path, monkeypatch):
        import nordpy.export as export_mod

        monkeypatch.setattr(export_mod, "EXPORT_DIR", tmp_path)

        from nordpy.models import Account

        data = [Account(accid=1, accno="123", type="ASK")]
        path = export_xlsx(data, "test_accounts")

        assert path.exists()
        assert path.suffix == ".xlsx"

    def test_bold_headers(self, tmp_path, monkeypatch):
        import nordpy.export as export_mod

        monkeypatch.setattr(export_mod, "EXPORT_DIR", tmp_path)

        from nordpy.models import Account
        from openpyxl import load_workbook

        data = [Account(accid=1, accno="123", type="ASK")]
        path = export_xlsx(data, "accounts")

        wb = load_workbook(path)
        ws = wb.active
        # Check header row is bold
        assert ws.cell(row=1, column=1).font.bold is True

    def test_number_formatting(self, tmp_path, monkeypatch):
        import nordpy.export as export_mod

        monkeypatch.setattr(export_mod, "EXPORT_DIR", tmp_path)
        from openpyxl import load_workbook

        h = Holding.model_validate(
            {
                "instrument": {"name": "X"},
                "qty": 10.0,
                "acq_price": {"value": 150.0, "currencyCode": "USD"},
                "market_value": {"value": 175.0, "currencyCode": "USD"},
            }
        )
        path = export_xlsx([h], "holdings")

        wb = load_workbook(path)
        ws = wb.active
        # Find a float cell and check its number format
        for row in ws.iter_rows(min_row=2, max_row=2):
            for cell in row:
                if isinstance(cell.value, float):
                    assert cell.number_format != "General"
                    break


# ── DuckDB export ──


class TestExportDuckDB:
    def test_creates_file(self, tmp_path, monkeypatch):
        import nordpy.export as export_mod

        monkeypatch.setattr(export_mod, "EXPORT_DIR", tmp_path)

        from nordpy.models import Account

        data = [Account(accid=1, accno="123", type="ASK")]
        path = export_duckdb(data, "test_accounts")

        assert path.exists()
        assert path.suffix == ".duckdb"

    def test_table_structure(self, tmp_path, monkeypatch):
        import duckdb

        import nordpy.export as export_mod

        monkeypatch.setattr(export_mod, "EXPORT_DIR", tmp_path)

        from nordpy.models import Account

        data = [
            Account(accid=1, accno="111", type="ASK"),
            Account(accid=2, accno="222", type="ISK"),
        ]
        path = export_duckdb(data, "accounts")

        con = duckdb.connect(str(path))
        result = con.execute("SELECT COUNT(*) FROM accounts").fetchone()
        assert result[0] == 2

        cols = con.execute("DESCRIBE accounts").fetchall()
        col_names = [c[0] for c in cols]
        assert "accid" in col_names
        assert "accno" in col_names
        con.close()

    def test_empty_data(self, tmp_path, monkeypatch):
        """DuckDB requires at least one column, so empty data raises an error."""
        import duckdb

        import nordpy.export as export_mod

        monkeypatch.setattr(export_mod, "EXPORT_DIR", tmp_path)

        with pytest.raises(duckdb.ParserException):
            export_duckdb([], "empty")
