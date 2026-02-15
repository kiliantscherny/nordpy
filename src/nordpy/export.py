"""Exporters — CSV, XLSX, and DuckDB export for Pydantic model data."""

from __future__ import annotations

import csv
from collections.abc import Sequence
from datetime import date, datetime
from pathlib import Path
from typing import Any

from pydantic import BaseModel

EXPORT_DIR = Path("exports")


def _ensure_export_dir() -> Path:
    EXPORT_DIR.mkdir(exist_ok=True)
    return EXPORT_DIR


def _timestamp() -> str:
    return datetime.now().strftime("%Y-%m-%d_%H%M%S")


def _model_to_flat_dict(item: BaseModel) -> dict[str, Any]:
    """Flatten a Pydantic model to a dict suitable for tabular export.

    For nested BaseModel fields that are None, expands them to sub-fields
    based on the field's annotation to ensure consistent column structure.
    """
    result: dict[str, Any] = {}
    model_fields = type(item).model_fields

    for field_name, value in item:
        if isinstance(value, BaseModel):
            # Nested model with a value - expand its fields
            for sub_name, sub_value in value:
                result[f"{field_name}_{sub_name}"] = sub_value
        elif value is None and field_name in model_fields:
            # Check if this None field is supposed to be a nested model
            field_info = model_fields[field_name]
            annotation = field_info.annotation
            # Handle Optional[SomeModel] by extracting the inner type
            nested_type = _get_nested_model_type(annotation)
            if nested_type is not None:
                # Expand None to sub-fields with None values
                for sub_field_name in nested_type.model_fields:
                    result[f"{field_name}_{sub_field_name}"] = None
            else:
                result[field_name] = value
        elif isinstance(value, date):
            result[field_name] = value.isoformat()
        else:
            result[field_name] = value
    return result


def _get_nested_model_type(annotation: Any) -> type[BaseModel] | None:
    """Extract a BaseModel type from an annotation like Optional[MoneyAmount]."""
    import types
    from typing import Union, get_args, get_origin

    origin = get_origin(annotation)

    # Handle Union types (e.g., MoneyAmount | None or Optional[MoneyAmount])
    if origin is Union or origin is types.UnionType:
        for arg in get_args(annotation):
            if isinstance(arg, type) and issubclass(arg, BaseModel):
                return arg

    # Handle direct BaseModel subclass
    if isinstance(annotation, type) and issubclass(annotation, BaseModel):
        return annotation

    return None


def _get_headers(data: Sequence[BaseModel]) -> list[str]:
    """Extract column headers from the first item."""
    if not data:
        return []
    return list(_model_to_flat_dict(data[0]).keys())


def _get_rows(data: Sequence[BaseModel]) -> list[list[Any]]:
    """Convert models to flat row lists."""
    return [list(_model_to_flat_dict(item).values()) for item in data]


# ── CSV Exporter ──


def export_csv(data: Sequence[BaseModel], entity_name: str) -> Path:
    """Export data to CSV format with column headers and timestamped filename."""
    _ensure_export_dir()
    filename = f"{entity_name}_{_timestamp()}.csv"
    path = EXPORT_DIR / filename

    headers = _get_headers(data)
    rows = _get_rows(data)

    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(rows)

    return path


# ── XLSX Exporter ──


def export_xlsx(data: Sequence[BaseModel], entity_name: str) -> Path:
    """Export data to Excel format with formatting and timestamped filename."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, numbers

    _ensure_export_dir()
    filename = f"{entity_name}_{_timestamp()}.xlsx"
    path = EXPORT_DIR / filename

    wb = Workbook()
    ws = wb.active
    ws.title = entity_name

    headers = _get_headers(data)
    rows = _get_rows(data)

    # Write header row with bold
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)

    # Write data rows
    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if isinstance(value, float):
                cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

    # Auto-adjust column widths
    for col_idx, header in enumerate(headers, 1):
        max_len = len(str(header))
        for row in rows:
            if col_idx - 1 < len(row):
                max_len = max(max_len, len(str(row[col_idx - 1])))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(
            max_len + 2, 50
        )

    wb.save(path)
    return path


# ── DuckDB Exporter ──


def export_duckdb(data: Sequence[BaseModel], entity_name: str) -> Path:
    """Export data to DuckDB format with structured table and timestamped filename."""
    import duckdb

    _ensure_export_dir()
    filename = f"{entity_name}_{_timestamp()}.duckdb"
    path = EXPORT_DIR / filename

    headers = _get_headers(data)
    rows = _get_rows(data)

    con = duckdb.connect(str(path))

    # Build column definitions
    col_defs = []
    if rows:
        for i, header in enumerate(headers):
            sample = rows[0][i] if i < len(rows[0]) else None
            sql_type = _python_to_sql_type(sample)
            safe_name = header.replace(" ", "_").replace("/", "_")
            col_defs.append(f'"{safe_name}" {sql_type}')
    else:
        for header in headers:
            safe_name = header.replace(" ", "_").replace("/", "_")
            col_defs.append(f'"{safe_name}" TEXT')

    table_name = entity_name.replace(" ", "_").replace("-", "_")
    create_sql = f'CREATE TABLE "{table_name}" ({", ".join(col_defs)})'
    con.execute(create_sql)

    if rows:
        placeholders = ", ".join(["?"] * len(headers))
        insert_sql = f'INSERT INTO "{table_name}" VALUES ({placeholders})'
        for row in rows:
            con.execute(insert_sql, row)

    con.close()
    return path


def _python_to_sql_type(value: object) -> str:
    """Map a Python value to a DuckDB SQL type."""
    if isinstance(value, float):
        return "DOUBLE"
    if isinstance(value, int):
        return "BIGINT"
    if isinstance(value, bool):
        return "BOOLEAN"
    return "TEXT"
